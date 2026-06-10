from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COLUNAS_EXPORT = [
    "cod_item_obra",
    "des_item_obra",
    "unidade_medida_inferida",
    "qtd_final_engenheiro",
    "val_unitario_final",
    "val_total_final",
]


def _formatar_planilha(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D9EAF7"))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for col in ws.columns:
        max_len = 12
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)) + 2, 60))
        ws.column_dimensions[col_letter].width = max_len


def _adicionar_dataframe(wb: Workbook, nome_aba: str, df: pd.DataFrame, colunas: list[str] | None = None) -> None:
    ws = wb.create_sheet(nome_aba)
    if df is None or df.empty:
        ws.append(["Sem itens"])
        return
    export_df = df.copy()
    if colunas:
        existentes = [c for c in colunas if c in export_df.columns]
        export_df = export_df[existentes]
    ws.append(list(export_df.columns))
    for _, row in export_df.iterrows():
        ws.append(list(row.values))
    _formatar_planilha(ws)


def gerar_excel_orcamento(
    resumo: dict[str, Any],
    materiais: pd.DataFrame,
    mao_de_obra: pd.DataFrame,
    taxas: pd.DataFrame,
    auditoria: pd.DataFrame,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Campo", "Valor"])
    for chave, valor in resumo.items():
        ws.append([chave, valor])
    _formatar_planilha(ws)

    _adicionar_dataframe(wb, "Materiais", materiais, COLUNAS_EXPORT)
    _adicionar_dataframe(wb, "Mao de obra", mao_de_obra, COLUNAS_EXPORT)
    _adicionar_dataframe(wb, "Taxas", taxas, COLUNAS_EXPORT)
    _adicionar_dataframe(wb, "Auditoria", auditoria, None)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
